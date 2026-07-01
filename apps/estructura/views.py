from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from apps.accounts.models import Empresa, User, UserEmpresa, Rol
from apps.actividades.models import TipoActividad, Actividad
from .models import Area, SubArea, UserSubArea, EmpresaArea
from .forms import AreaForm, SubAreaForm, EmpresaForm


def master_required(view_func):
    def wrapper(request, *args, **kwargs):
        if request.user.rol.nombre != "Master":
            return redirect("root")
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
@master_required
def empresa_list(request):
    q = request.GET.get("q", "")
    empresas = Empresa.objects.filter(activo=True)
    if q:
        empresas = empresas.filter(
            Q(nombre__icontains=q) | Q(nit__icontains=q) | Q(telefono__icontains=q)
        )
    paginator = Paginator(empresas, 10)
    page = request.GET.get("page", 1)
    empresas_page = paginator.get_page(page)
    return render(request, "estructura/empresa_list.html", {"empresas": empresas_page, "page_obj": empresas_page, "q": q})


@login_required
@master_required
def empresa_create(request):
    if request.method == "POST":
        form = EmpresaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Empresa creada exitosamente.")
            return redirect("estructura:empresa_list")
    else:
        form = EmpresaForm()
    return render(request, "estructura/empresa_form.html", {"form": form, "title": "Crear Empresa"})


@login_required
@master_required
def empresa_edit(request, pk):
    empresa = get_object_or_404(Empresa, pk=pk, activo=True)
    if request.method == "POST":
        form = EmpresaForm(request.POST, request.FILES, instance=empresa)
        if form.is_valid():
            form.save()
            messages.success(request, "Empresa actualizada exitosamente.")
            return redirect("estructura:empresa_list")
    else:
        form = EmpresaForm(instance=empresa)
    return render(request, "estructura/empresa_form.html", {"form": form, "title": "Editar Empresa"})


@login_required
@master_required
def empresa_delete(request, pk):
    empresa = get_object_or_404(Empresa, pk=pk)
    empresa.activo = False
    empresa.save()
    messages.success(request, "Empresa inactivada exitosamente.")
    return redirect("estructura:empresa_list")


@login_required
@master_required
def area_list(request):
    q = request.GET.get("q", "")
    empresa_filter = request.GET.get("empresa", "")
    areas = Area.objects.filter(activo=True)
    if q:
        areas = areas.filter(Q(nombre__icontains=q) | Q(descripcion__icontains=q))
    paginator = Paginator(areas, 10)
    page = request.GET.get("page", 1)
    areas_page = paginator.get_page(page)
    empresas = Empresa.objects.filter(activo=True)
    return render(request, "estructura/area_list.html", {"areas": areas_page, "page_obj": areas_page, "empresas": empresas, "q": q, "empresa_filter": empresa_filter})


@login_required
@master_required
def area_create(request):
    if request.method == "POST":
        form = AreaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Area creada exitosamente.")
            return redirect("estructura:area_list")
    else:
        form = AreaForm()
    return render(request, "estructura/area_form.html", {"form": form, "title": "Crear Area"})


@login_required
@master_required
def area_edit(request, pk):
    area = get_object_or_404(Area, pk=pk, activo=True)
    if request.method == "POST":
        form = AreaForm(request.POST, instance=area)
        if form.is_valid():
            form.save()
            messages.success(request, "Area actualizada exitosamente.")
            return redirect("estructura:area_list")
    else:
        form = AreaForm(instance=area)
    return render(request, "estructura/area_form.html", {"form": form, "title": "Editar Area", "area": area})


@login_required
@master_required
def area_delete(request, pk):
    area = get_object_or_404(Area, pk=pk)
    tiene_activo = Actividad.objects.filter(
        subarea__area=area, activo=True, asignaciones__activo=True
    ).exclude(asignaciones__estado__in=["Pendiente", "Cancelada"]).exists()
    if tiene_activo:
        messages.error(request, "No se puede inactivar el area porque tiene actividades en curso o pausadas.")
        return redirect("estructura:area_list")
    for subarea in SubArea.objects.filter(area=area, activo=True):
        _inactivar_subarea(subarea)
    area.activo = False
    area.save()
    messages.success(request, "Area inactivada exitosamente.")
    return redirect("estructura:area_list")


@login_required
@master_required
def subarea_list(request):
    q = request.GET.get("q", "")
    area_filter = request.GET.get("area", "")
    empresa_filter = request.GET.get("empresa", "")
    subareas = SubArea.objects.filter(activo=True)
    if q:
        subareas = subareas.filter(Q(nombre__icontains=q) | Q(descripcion__icontains=q))
    if area_filter:
        subareas = subareas.filter(area_id=area_filter)
    paginator = Paginator(subareas, 10)
    page = request.GET.get("page", 1)
    subareas_page = paginator.get_page(page)
    empresas = Empresa.objects.filter(activo=True)
    areas = Area.objects.filter(activo=True)
    return render(request, "estructura/subarea_list.html", {
        "subareas": subareas_page, "page_obj": subareas_page, "empresas": empresas, "areas": areas,
        "q": q, "area_filter": area_filter
    })


@login_required
@master_required
def subarea_create(request):
    usuarios = User.objects.filter(activo=True, is_active=True).exclude(rol__nombre="Master")
    if request.method == "POST":
        form = SubAreaForm(request.POST)
        if form.is_valid():
            subarea = form.save()
            user_ids = request.POST.getlist("users")
            for uid in user_ids:
                UserSubArea.objects.get_or_create(user_id=uid, subarea=subarea, defaults={"activo": True})
            messages.success(request, "Subarea creada con usuarios asignados.")
            return redirect("estructura:subarea_list")
    else:
        form = SubAreaForm()
    return render(request, "estructura/subarea_form.html", {"form": form, "title": "Crear SubArea", "usuarios": usuarios})


@login_required
@master_required
def subarea_edit(request, pk):
    subarea = get_object_or_404(SubArea, pk=pk, activo=True)
    usuarios = User.objects.filter(activo=True, is_active=True).exclude(rol__nombre="Master")
    users_asignados = list(UserSubArea.objects.filter(subarea=subarea, activo=True).values_list("user_id", flat=True))

    if request.method == "POST":
        form = SubAreaForm(request.POST, instance=subarea)
        if form.is_valid():
            form.save()
            user_ids = set(request.POST.getlist("users", []))
            current_users = set(UserSubArea.objects.filter(subarea=subarea, activo=True).values_list("user_id", flat=True))

            to_add = user_ids - current_users
            to_remove = current_users - user_ids
            for uid in to_add:
                UserSubArea.objects.get_or_create(user_id=uid, subarea=subarea, defaults={"activo": True})
            UserSubArea.objects.filter(subarea=subarea, user_id__in=to_remove).update(activo=False)

            messages.success(request, "Subarea actualizada con usuarios.")
            return redirect("estructura:subarea_list")
    else:
        form = SubAreaForm(instance=subarea)

    return render(request, "estructura/subarea_form.html", {
        "form": form, "title": "Editar SubArea", "subarea": subarea,
        "usuarios": usuarios, "users_asignados": users_asignados,
    })


def _inactivar_subarea(subarea):
    from apps.planificacion.models import Planificacion, PlanificacionDetalle
    from apps.gestion.models import AsignacionActividad
    from apps.actividades.models import Actividad
    for p in Planificacion.objects.filter(subarea=subarea, activo=True):
        for detalle in PlanificacionDetalle.objects.filter(planificacion=p, activo=True):
            AsignacionActividad.objects.filter(planificacion_detalle=detalle, activo=True).update(activo=False, estado="Cancelada")
            detalle.activo = False
            detalle.save()
        p.activo = False
        p.save()
    Actividad.objects.filter(subarea=subarea, activo=True).update(activo=False)
    subarea.activo = False
    subarea.save()


@login_required
@master_required
def subarea_delete(request, pk):
    subarea = get_object_or_404(SubArea, pk=pk)
    tiene_activo = Actividad.objects.filter(
        subarea=subarea, activo=True, asignaciones__activo=True
    ).exclude(asignaciones__estado__in=["Pendiente", "Cancelada"]).exists()
    if tiene_activo:
        messages.error(request, "No se puede inactivar la subarea porque tiene actividades en curso o pausadas.")
        return redirect("estructura:subarea_list")
    _inactivar_subarea(subarea)
    messages.success(request, "Subarea inactivada exitosamente.")
    return redirect("estructura:subarea_list")


@login_required
@master_required
def subarea_usuarios(request, pk):
    subarea = get_object_or_404(SubArea, pk=pk, activo=True)
    usuarios_asignados = UserSubArea.objects.filter(subarea=subarea, activo=True).select_related("user")
    usuarios_disponibles = User.objects.filter(activo=True, is_active=True).exclude(
        id__in=usuarios_asignados.values_list("user_id", flat=True)
    )
    if request.method == "POST":
        user_id = request.POST.get("user_id")
        if user_id:
            UserSubArea.objects.get_or_create(
                user_id=user_id, subarea=subarea, defaults={"activo": True}
            )
            messages.success(request, "Usuario asignado a la subarea.")
            return redirect("estructura:subarea_usuarios", pk=pk)
    return render(request, "estructura/subarea_usuarios.html", {
        "subarea": subarea,
        "usuarios_asignados": usuarios_asignados,
        "usuarios_disponibles": usuarios_disponibles,
    })


@login_required
@master_required
def subarea_usuario_remove(request, pk, user_pk):
    usa = get_object_or_404(UserSubArea, subarea_id=pk, user_id=user_pk)
    usa.activo = False
    usa.save()
    messages.success(request, "Usuario removido de la subarea.")
    return redirect("estructura:subarea_usuarios", pk=pk)


# --- API para selects dinamicos ---
@login_required
def api_buscar(request, modelo):
    if not request.user.is_authenticated:
        return JsonResponse([], safe=False)

    # Para Admins (no Master), limitar subareas a las suyas
    is_admin_scoped = request.user.rol.nombre not in ["Master"] and modelo in ("subarea", "actividad", "tipo_actividad", "user")

    q = request.GET.get("q", "")
    results = []
    if modelo == "empresa":
        qs = Empresa.objects.filter(activo=True)
        if q:
            qs = qs.filter(Q(nombre__icontains=q) | Q(nit__icontains=q))
        for o in qs[:20]:
            results.append({"id": o.pk, "text": f"{o.nombre} ({o.nit})", "label": o.nombre})
    elif modelo == "area":
        qs = Area.objects.filter(activo=True)
        if q:
            qs = qs.filter(Q(nombre__icontains=q))
        for o in qs[:20]:
            results.append({"id": o.pk, "text": o.nombre, "label": o.nombre})
    elif modelo == "subarea":
        qs = SubArea.objects.filter(activo=True)
        if is_admin_scoped:
            qs = qs.filter(usuarios__user=request.user, usuarios__activo=True)
        if q:
            qs = qs.filter(Q(nombre__icontains=q))
        area_id = request.GET.get("area_id")
        if area_id:
            qs = qs.filter(area_id=area_id)
        for o in qs[:20]:
            results.append({"id": o.pk, "text": f"{o.nombre} ({o.area.nombre})", "label": o.nombre})
    elif modelo == "tipo_actividad":
        from apps.actividades.models import TipoActividad
        qs = TipoActividad.objects.filter(activo=True)
        if is_admin_scoped:
            qs = qs.filter(subarea__usuarios__user=request.user, subarea__usuarios__activo=True)
        if q:
            qs = qs.filter(Q(nombre__icontains=q))
        subarea_id = request.GET.get("subarea_id")
        if subarea_id:
            qs = qs.filter(subarea_id=subarea_id)
        for o in qs[:20]:
            results.append({"id": o.pk, "text": f"{o.nombre} ({o.subarea.nombre})", "label": o.nombre})
    elif modelo == "actividad":
        from apps.actividades.models import Actividad
        qs = Actividad.objects.filter(activo=True).select_related("tipo_actividad")
        if is_admin_scoped:
            qs = qs.filter(subarea__usuarios__user=request.user, subarea__usuarios__activo=True)
        if q:
            qs = qs.filter(Q(nombre__icontains=q))
        subarea_id = request.GET.get("subarea_id")
        if subarea_id:
            qs = qs.filter(subarea_id=subarea_id)
        tipo_nombre = request.GET.get("tipo_nombre")
        if tipo_nombre:
            qs = qs.filter(tipo_actividad__nombre__iexact=tipo_nombre)
        tipo_id = request.GET.get("tipo_id")
        if tipo_id:
            qs = qs.filter(tipo_actividad_id=tipo_id)
        if request.GET.get("es_flash"):
            qs = qs.filter(tipo_actividad__es_flash=True)
        limite_default = int(request.GET.get("limite", 0))
        limite = limite_default if limite_default else (20 if q else 50 if request.GET.get("es_flash") else 5)
        for o in qs[:limite]:
            results.append({
                "id": o.pk,
                "text": f"{o.nombre} ({o.tipo_actividad.nombre})",
                "label": o.nombre,
                "tipo_nombre": o.tipo_actividad.nombre,
                "subarea_id": o.subarea_id,
                "requiere_fecha_limite": o.tipo_actividad.requiere_fecha_limite,
            })
    elif modelo == "user":
        qs = User.objects.filter(activo=True, is_active=True).filter(
            Q(rol__nombre="Usuario") | Q(roles_adicionales__nombre="Usuario")
        ).distinct()
        subarea_id = request.GET.get("subarea_id")
        if is_admin_scoped and not subarea_id:
            qs = qs.filter(subareas__subarea__usuarios__user=request.user, subareas__activo=True)
        if q:
            qs = qs.filter(Q(nombre__icontains=q) | Q(apellido__icontains=q) | Q(cedula__icontains=q))
        if subarea_id:
            qs = qs.filter(subareas__subarea_id=subarea_id, subareas__activo=True)
        # Ordenar por mas reciente asignacion
        from django.db.models import Max
        qs = qs.annotate(ultima_asignacion=Max("asignaciones__fecha_asignacion")).order_by("-ultima_asignacion", "nombre")
        limite_default = int(request.GET.get("limite", 0))
        limite = limite_default if limite_default else (20 if q else 10)
        for o in qs.distinct()[:limite]:
            results.append({"id": o.pk, "text": f"{o.get_full_name()} ({o.cedula})", "label": o.get_full_name()})
    return JsonResponse(results, safe=False)


def _estilo_header(ws, headers):
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center")


@login_required
@master_required
def importar_exportar(request):
    return render(request, "estructura/importar.html")


@login_required
@master_required
def descargar_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "AreasSubAreas"

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    headers = ["nombre_area", "nombre_subarea"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center")
    ws.append(["", ""])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="areas_subareas.xlsx"'
    wb.save(resp)
    return resp


@login_required
@master_required
def importar_datos(request):
    if request.method != "POST" or not request.FILES.get("archivo"):
        messages.error(request, "Selecciona un archivo Excel.")
        return redirect("estructura:importar_exportar")

    archivo = request.FILES["archivo"]
    try:
        wb = load_workbook(archivo, read_only=True)
        ws = wb.active
    except Exception:
        messages.error(request, "El archivo no es un Excel valido.")
        return redirect("estructura:importar_exportar")

    errores = []
    creados = {"areas": 0, "subareas": 0}

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        nombre_a = (str(row[0] or "")).strip()
        nombre_s = (str(row[1] or "")).strip()
        if not nombre_a or not nombre_s:
            errores.append(f"Fila {i}: nombre_area y nombre_subarea requeridos"); continue

        try:
            area, created_a = Area.objects.get_or_create(
                nombre=nombre_a,
                defaults={"codigo": None}
            )
            if created_a: creados["areas"] += 1
            sub, created_s = SubArea.objects.get_or_create(
                nombre=nombre_s, area=area,
                defaults={"codigo": None}
            )
            if created_s: creados["subareas"] += 1
        except Exception as e:
            errores.append(f"Fila {i}: {str(e)[:80]}")

    wb.close()
    total = creados["areas"] + creados["subareas"]
    if total > 0:
        messages.success(request, f"Creadas: {creados['areas']} area(s) y {creados['subareas']} subarea(s).")
    for err in errores[:5]:
        messages.warning(request, err)
    if not total and not errores:
        messages.warning(request, "No se creo nada. Verifica el formato del archivo.")
    return redirect("estructura:importar_exportar")


@login_required
@master_required
def importar_usuarios(request):
    return render(request, "estructura/importar_usuarios.html")


@login_required
@master_required
def descargar_template_usuarios(request):
    from apps.accounts.models import Rol
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    headers = ["cedula", "area_codigo", "subarea_codigo"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center")

    ws.append(["1234567890", "AR001", "SB001"])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="usuarios_matricula.xlsx"'
    wb.save(resp)
    return resp


@login_required
@master_required
def importar_usuarios_datos(request):
    if request.method != "POST" or not request.FILES.get("archivo"):
        messages.error(request, "Selecciona un archivo Excel.")
        return redirect("estructura:importar_usuarios")

    archivo = request.FILES["archivo"]
    try:
        wb = load_workbook(archivo, read_only=True)
        ws = wb.active
    except Exception:
        messages.error(request, "El archivo no es un Excel valido.")
        return redirect("estructura:importar_usuarios")

    errores = []
    matriculados = 0

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        cedula = (str(row[0] or "")).strip()
        area_cod = (str(row[1] or "")).strip()
        subarea_cod = (str(row[2] or "")).strip()

        if not cedula or not area_cod or not subarea_cod:
            errores.append(f"Fila {i}: cedula, area_codigo y subarea_codigo requeridos")
            continue

        user = User.objects.filter(cedula=cedula, activo=True).first()
        if not user:
            errores.append(f"Fila {i}: cedula '{cedula}' no existe en el sistema")
            continue

        area = Area.objects.filter(codigo=area_cod, activo=True).first()
        if not area:
            errores.append(f"Fila {i}: area_codigo '{area_cod}' no existe")
            continue

        subarea = SubArea.objects.filter(codigo=subarea_cod, area=area, activo=True).first()
        if not subarea:
            errores.append(f"Fila {i}: subarea_codigo '{subarea_cod}' no existe para area '{area_cod}'")
            continue

        try:
            obj, created = UserSubArea.objects.update_or_create(
                user=user, subarea=subarea,
                defaults={"activo": True}
            )
            if created:
                matriculados += 1
            else:
                matriculados += 1  # reactivada
        except Exception as e:
            errores.append(f"Fila {i}: {str(e)[:80]}")

    wb.close()
    if matriculados > 0:
        messages.success(request, f"{matriculados} usuario(s) matriculados exitosamente.")
    for err in errores[:5]:
        messages.warning(request, err)
    if not matriculados and not errores:
        messages.warning(request, "No se matricularon usuarios. Verifica el formato del archivo.")
    return redirect("estructura:importar_usuarios")
