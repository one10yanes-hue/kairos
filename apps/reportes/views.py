from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q, Count
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from apps.gestion.models import AsignacionActividad, RegistroTiempo
from apps.accounts.models import User
from apps.estructura.models import Area, SubArea, UserSubArea, EmpresaArea
from apps.actividades.models import Actividad


def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not (request.user.rol.nombre in ["Admin", "Master"] or
                request.user.roles_adicionales.filter(nombre__in=["Admin", "Master"]).exists()):
            from django.shortcuts import redirect
            return redirect("root")
        return view_func(request, *args, **kwargs)
    return wrapper


def get_admin_subareas(user):
    if user.rol.nombre == "Master" or user.roles_adicionales.filter(nombre="Master").exists():
        return SubArea.objects.filter(activo=True)
    # Check original primary role from DB (in case role switcher changed user.rol)
    db_rol = User.objects.filter(pk=user.pk).values_list('rol__nombre', flat=True).first()
    if db_rol == "Master":
        return SubArea.objects.filter(activo=True)
    return SubArea.objects.filter(usuarios__user=user, activo=True)


HF = Font(bold=True, color="FFFFFF", size=11)
HFILL = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
AL = Alignment(horizontal="center", vertical="center", wrap_text=True)
TB = Border(
    left=Side(style="thin", color="e2e8f0"),
    right=Side(style="thin", color="e2e8f0"),
    top=Side(style="thin", color="e2e8f0"),
    bottom=Side(style="thin", color="e2e8f0"),
)


def _style_header(ws, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HF; c.fill = HFILL; c.alignment = AL; c.border = TB


def _auto_width(ws):
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 3, 45)


@login_required
@admin_required
def reporte_list(request):
    subareas = get_admin_subareas(request.user)
    empresas = EmpresaArea.objects.filter(area__subareas__in=subareas).values_list("empresa__nombre", flat=True).distinct()
    areas = Area.objects.filter(subareas__in=subareas)
    users = User.objects.filter(
        id__in=UserSubArea.objects.filter(subarea__in=subareas, activo=True).values_list("user_id", flat=True),
        activo=True
    ).exclude(rol__nombre="Master")
    db_es_master = User.objects.filter(pk=request.user.pk, rol__nombre="Master").exists()
    if not (request.user.rol.nombre == "Master" or request.user.roles_adicionales.filter(nombre="Master").exists() or db_es_master):
        users = users.filter(
            Q(rol__nombre="Usuario") | Q(roles_adicionales__nombre="Usuario")
        ).distinct()

    qs = AsignacionActividad.objects.filter(actividad__subarea__in=subareas, activo=True)
    return render(request, "reportes/reporte_list.html", {
        "subareas": subareas, "empresas": empresas, "areas": areas, "users": users,
        "total_act": qs.count(),
        "total_fin": qs.filter(estado="Finalizada").count(),
        "total_curso": qs.filter(estado="EnCurso").count(),
        "total_pausa": qs.filter(estado="Pausada").count(),
        "total_pen": qs.filter(estado="Pendiente").count(),
    })


@login_required
@admin_required
def exportar_completo(request):
    subareas = get_admin_subareas(request.user)
    q = request.GET.get("q", "")
    empresa_nombre = request.GET.get("empresa", "")
    area_id = request.GET.get("area", "")
    subarea_id = request.GET.get("subarea", "")
    user_id = request.GET.get("user", "")
    estado = request.GET.get("estado", "")
    fecha_desde = request.GET.get("fecha_desde", "")
    fecha_hasta = request.GET.get("fecha_hasta", "")

    asignaciones = AsignacionActividad.objects.filter(
        actividad__subarea__in=subareas, activo=True
    )
    asignaciones_tiempo = AsignacionActividad.objects.filter(
        actividad__subarea__in=subareas
    )
    asignaciones = asignaciones.select_related(
        "user", "actividad", "actividad__tipo_actividad",
        "actividad__subarea", "actividad__subarea__area",
        "planificacion_detalle__planificacion__admin"
    ).prefetch_related("registros")

    if q:
        asignaciones = asignaciones.filter(
            Q(actividad__nombre__icontains=q) |
            Q(user__nombre__icontains=q) |
            Q(user__apellido__icontains=q)
        )
        asignaciones_tiempo = asignaciones_tiempo.filter(
            Q(actividad__nombre__icontains=q) |
            Q(user__nombre__icontains=q) |
            Q(user__apellido__icontains=q)
        )
    if area_id:
        asignaciones = asignaciones.filter(actividad__subarea__area_id=area_id)
        asignaciones_tiempo = asignaciones_tiempo.filter(actividad__subarea__area_id=area_id)
    if subarea_id:
        asignaciones = asignaciones.filter(actividad__subarea_id=subarea_id)
        asignaciones_tiempo = asignaciones_tiempo.filter(actividad__subarea_id=subarea_id)
    if user_id:
        asignaciones = asignaciones.filter(user_id=user_id)
        asignaciones_tiempo = asignaciones_tiempo.filter(user_id=user_id)
    if estado:
        asignaciones = asignaciones.filter(estado=estado)
        asignaciones_tiempo = asignaciones_tiempo.filter(estado=estado)
    if fecha_desde:
        asignaciones = asignaciones.filter(fecha_asignacion__date__gte=fecha_desde)
        asignaciones_tiempo = asignaciones_tiempo.filter(fecha_asignacion__date__gte=fecha_desde)
    if fecha_hasta:
        asignaciones = asignaciones.filter(fecha_asignacion__date__lte=fecha_hasta)
        asignaciones_tiempo = asignaciones_tiempo.filter(fecha_asignacion__date__lte=fecha_hasta)

    wb = Workbook()

    # ---- SHEET 1: Actividades ----
    ws = wb.active; ws.title = "Actividades"
    headers = [
        "ID Asig", "Cod. Empresa", "Empresa",
        "Cod. Area", "Area",
        "Cod. SubArea", "SubArea",
        "Cod. Actividad", "Actividad",
        "Cod. Tipo", "Tipo",
        "Usuario", "Cedula", "Email",
        "Estado", "Origen", "Asignado por",
        "Fecha Asignacion", "Fecha Vencimiento",
        "Tiempo Efectivo",
        "Fecha Primer Inicio", "Fecha Ultimo Evento",
        "Nro Actividad", "Total Pausas",
    ]
    _style_header(ws, headers)

    for row_num, a in enumerate(asignaciones, 2):
        r = list(a.registros.filter(activo=True).order_by("fecha_hora"))
        tz = timezone.get_current_timezone()
        inicio = r[0].fecha_hora.astimezone(tz).strftime("%Y-%m-%d %H:%M") if r and r[0].evento == "Inicio" else ""
        ultimo = r[-1].fecha_hora.astimezone(tz).strftime("%Y-%m-%d %H:%M") if r else ""
        nros = ", ".join(filter(None, (x.nro_actividad for x in r if x.nro_actividad)))
        pausas = sum(1 for x in r if x.evento == "Pausa")

        sub = a.actividad.subarea
        pd = a.planificacion_detalle
        venc = pd.fecha_vencimiento.astimezone(tz).strftime("%Y-%m-%d") if pd and pd.fecha_vencimiento and a.actividad.tipo_actividad.requiere_fecha_limite else "-"
        row = [
            a.pk,
            "", "",
            sub.area.codigo, sub.area.nombre,
            sub.codigo, sub.nombre,
            a.actividad.codigo, a.actividad.nombre,
            a.actividad.tipo_actividad.codigo, a.actividad.tipo_actividad.nombre,
            a.user.get_full_name(), a.user.cedula, a.user.email or "",
            a.get_estado_display(), a.origen or "Manual",
            a.origen_user.get_full_name() if a.origen_user else "-",
            a.fecha_asignacion.astimezone(tz).strftime("%Y-%m-%d %H:%M") if a.fecha_asignacion else "",
            venc,
            a.tiempo_formateado(),
            inicio, ultimo,
            nros, pausas,
        ]
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=c, value=v)
            cell.border = TB
            if c in (2, 11, 15, 16, 17):
                cell.alignment = Alignment(horizontal="left")
    _auto_width(ws)

    # ---- SHEET 2: Registros de Tiempo ----
    ws2 = wb.create_sheet("Registros Tiempo")
    headers2 = [
        "ID Asig", "ID Evento", "Empresa", "Area", "SubArea", "Actividad", "Tipo",
        "Usuario", "Cedula",
        "Evento", "Fecha/Hora", "Motivo Pausa", "Comentario", "Nro Actividad",
    ]
    _style_header(ws2, headers2)

    registros = RegistroTiempo.objects.filter(
        asignacion__actividad__subarea__in=subareas, activo=True
    ).select_related(
        "asignacion__user", "asignacion__actividad",
    ).order_by("-fecha_hora")

    if fecha_desde:
        registros = registros.filter(fecha_hora__date__gte=fecha_desde)
    if fecha_hasta:
        registros = registros.filter(fecha_hora__date__lte=fecha_hasta)
    if user_id:
        registros = registros.filter(asignacion__user_id=user_id)
    if subarea_id:
        registros = registros.filter(asignacion__actividad__subarea_id=subarea_id)

    tz2 = timezone.get_current_timezone()
    for row_num, r in enumerate(registros, 2):
        sub = r.asignacion.actividad.subarea
        row = [
            r.asignacion_id, r.pk,
            sub.area.nombre, sub.area.nombre, sub.nombre,
            r.asignacion.actividad.nombre,
            r.asignacion.actividad.tipo_actividad.nombre,
            r.asignacion.user.get_full_name(), r.asignacion.user.cedula,
            r.get_evento_display(),
            r.fecha_hora.astimezone(tz2).strftime("%Y-%m-%d %H:%M:%S"),
            r.motivo_pausa or "", r.comentario or "", r.nro_actividad or "",
        ]
        for c, v in enumerate(row, 1):
            cell = ws2.cell(row=row_num, column=c, value=v)
            cell.border = TB
    _auto_width(ws2)

    # ---- SHEET 3: Resumen por Usuario ----
    ws3 = wb.create_sheet("Resumen por Usuario")
    headers3 = [
        "Usuario", "Empresa", "Total Asignaciones",
        "Finalizadas", "En Curso", "Pausadas", "Pendientes",
        "Tiempo Total", "Nro Actividad Total",
    ]
    _style_header(ws3, headers3)

    for row_num, a in enumerate(asignaciones_tiempo.values("user__pk").annotate(total=Count("id")), 2):
        pk = a["user__pk"]
        asign_user = asignaciones_tiempo.filter(user__pk=pk)
        user = asign_user.first().user
        empresas = set(
            usa.subarea.area.nombre
            for usa in UserSubArea.objects.filter(user=user, activo=True).select_related("subarea__area")
        )
        total_tiempo = sum(x.tiempo_efectivo() for x in asign_user)
        row = [
            user.get_full_name(),
            ", ".join(empresas),
            a["total"],
            asign_user.filter(estado="Finalizada").count(),
            asign_user.filter(estado="EnCurso").count(),
            asign_user.filter(estado="Pausada").count(),
            asign_user.filter(estado="Pendiente").count(),
            f"{int(total_tiempo // 3600):02d}:{int((total_tiempo % 3600) // 60):02d}",
            sum(
                int(x.nro_actividad) for x in
                RegistroTiempo.objects.filter(
                    asignacion__in=asign_user, nro_actividad__isnull=False, activo=True
                ) if x.nro_actividad.isdigit()
            ),
        ]
        for c, v in enumerate(row, 1):
            ws3.cell(row=row_num, column=c, value=v).border = TB
    _auto_width(ws3)

    # ---- SHEET 4: Tiempos Muertos ----
    from apps.gestion.models import TiempoInactividad
    ws4 = wb.create_sheet("Tiempos Inactividad")
    headers4 = ["ID", "Usuario", "Cedula", "Fecha", "Inicio", "Fin", "Duracion (min)", "Estado"]
    _style_header(ws4, headers4)
    tiempos_qs = TiempoInactividad.objects.all().select_related("user").order_by("-fecha", "-inicio")
    if fecha_desde:
        tiempos_qs = tiempos_qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        tiempos_qs = tiempos_qs.filter(fecha__lte=fecha_hasta)
    if user_id:
        tiempos_qs = tiempos_qs.filter(user_id=user_id)
    for row_num, tm in enumerate(tiempos_qs, 2):
        mins = int(tm.duracion_segundos // 60) if tm.duracion_segundos > 0 else 0
        inicio_str = tm.inicio.strftime("%Y-%m-%d %H:%M") if tm.inicio else "-"
        fin_str = tm.fin.strftime("%Y-%m-%d %H:%M") if tm.fin else "Abierto"
        row = [tm.pk, tm.user.get_full_name(), tm.user.cedula, str(tm.fecha), inicio_str, fin_str, mins, "Cerrado" if tm.fin else "Abierto"]
        for c, v in enumerate(row, 1):
            ws4.cell(row=row_num, column=c, value=v).border = TB
    _auto_width(ws4)

    # ---- SHEET 5: Proyectos ----
    from apps.proyectos.models import Proyecto
    ws5 = wb.create_sheet("Proyectos")
    headers5 = ["Codigo", "Nombre", "Estado", "Manager", "Tareas Total", "Pendientes", "En Curso", "Finalizadas", "Inc. Abiertas", "Inc. Cerradas", "Avance %", "F. Inicio", "F. Fin Est."]
    _style_header(ws5, headers5)
    proyectos_qs = Proyecto.objects.filter(activo=True).select_related("manager")
    for row_num, p in enumerate(proyectos_qs, 2):
        t_tot = p.tareas.filter(activo=True).count()
        t_pen = p.tareas.filter(activo=True, estado="pendiente").count()
        t_cur = p.tareas.filter(activo=True, estado__in=["en_curso","pausada"]).count()
        t_fin = p.tareas.filter(activo=True, estado="finalizada").count()
        i_ab = p.incidencias.filter(activo=True, estado__in=["abierta","triaged","en_progreso"]).count()
        i_ce = p.incidencias.filter(activo=True, estado="cerrada").count()
        row = [p.codigo, p.nombre, p.get_estado_display(), p.manager.get_full_name(),
               t_tot, t_pen, t_cur, t_fin, i_ab, i_ce, f"{p.avance}%",
               str(p.fecha_inicio) if p.fecha_inicio else "-", str(p.fecha_fin_estimada) if p.fecha_fin_estimada else "-"]
        for c, v in enumerate(row, 1):
            ws5.cell(row=row_num, column=c, value=v).border = TB
    _auto_width(ws5)

    # ---- SHEET 6: Tareas por Proyecto ----
    from apps.proyectos.models import Tarea
    ws6 = wb.create_sheet("Tareas")
    headers6 = ["Codigo", "Titulo", "Proyecto", "Tipo", "Estado", "Asignado", "Sprint", "F. Creacion"]
    _style_header(ws6, headers6)
    tareas_qs = Tarea.objects.filter(activo=True).select_related("proyecto", "asignado_a", "sprint")
    for row_num, t in enumerate(tareas_qs, 2):
        row = [t.codigo, t.titulo, t.proyecto.codigo, t.get_tipo_display(), t.get_estado_display(),
               t.asignado_a.get_full_name() if t.asignado_a else "-", t.sprint.nombre if t.sprint else "-",
               t.fecha_creacion.strftime("%Y-%m-%d")]
        for c, v in enumerate(row, 1):
            ws6.cell(row=row_num, column=c, value=v).border = TB
    _auto_width(ws6)

    now = timezone.now().strftime("%Y%m%d_%H%M")
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="reporte_completo_{now}.xlsx"'
    wb.save(resp)
    return resp
